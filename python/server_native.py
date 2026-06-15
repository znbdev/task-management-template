#!/usr/bin/env python3
"""HTTP server (stdlib only, no Flask) for Task Management Board."""

import io
import json
import os
import re
import uuid
from datetime import datetime
from http.server import HTTPServer, BaseHTTPRequestHandler
from typing import Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from models import Task, WBSItem, DailyLogItem, RiskItem, ChangeHistoryItem
from models import STATUSES, PRIORITIES, TASK_TYPES, PHASES
from storage import load_tasks, save_tasks, create_task, update_task, delete_task, get_task

HOST = "0.0.0.0"
PORT = 5800

TEMPLATE_DIR = os.path.join(os.path.dirname(__file__), "templates")


def _task_to_dict(t: Task) -> dict:
    d = t.to_dict()
    d["total_estimated"] = t.total_estimated
    d["total_actual"] = t.total_actual
    d["status_icon"] = t.status_icon
    d["priority_icon"] = t.priority_icon
    return d


def _dict_to_task(data: dict) -> Task:
    def _parse_list(key, cls):
        items = data.get(key, [])
        return [cls(**item) if isinstance(item, dict) else item for item in items]

    return Task(
        id=data.get("id", ""),
        title=data.get("title", ""),
        description=data.get("description", ""),
        status=data.get("status", STATUSES[0]),
        priority=data.get("priority", PRIORITIES[2]),
        task_type=data.get("task_type", TASK_TYPES[0]),
        owner=data.get("owner", ""),
        requester=data.get("requester", ""),
        start_date=data.get("start_date", ""),
        due_date=data.get("due_date", ""),
        core_objective=data.get("core_objective", ""),
        deliverables=data.get("deliverables", []),
        success_criteria=data.get("success_criteria", []),
        wbs_items=_parse_list("wbs_items", WBSItem),
        risks=[RiskItem(**r) if isinstance(r, dict) else r for r in data.get("risks", [])],
        external_dependencies=data.get("external_dependencies", []),
        buffer_hours=float(data.get("buffer_hours", 0)),
        daily_logs=_parse_list("daily_logs", DailyLogItem),
        resources=data.get("resources", []),
        change_history=[ChangeHistoryItem(**h) if isinstance(h, dict) else h for h in data.get("change_history", [])],
        tags=data.get("tags", []),
        created_at=data.get("created_at", ""),
        updated_at=data.get("updated_at", ""),
    )


class TaskHandler(BaseHTTPRequestHandler):

    def _send_excel(self):
        tasks = load_tasks()
        wb = Workbook()
        wb.remove(wb.active)

        HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
        HEADER_FILL = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        SECTION_FONT = Font(bold=True, size=11, color="2C3E50")
        SECTION_FILL = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")
        THIN_BORDER = Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC"),
        )

        STATUS_ICONS_DICT = dict(zip(STATUSES, ["🟩", "🟨", "🟦", "⬜", "❌"]))
        PRIORITY_ICONS_DICT = dict(zip(PRIORITIES, ["🔴", "🟠", "🟡", "🔵"]))

        for task in tasks:
            sheet_name = task.title[:31] if task.title else "無題"
            ws = wb.create_sheet(title=sheet_name)

            r = 1

            def write_section(title, headers, rows, col_widths=None):
                nonlocal r
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
                cell = ws.cell(row=r, column=1, value=title)
                cell.font = SECTION_FONT
                cell.fill = SECTION_FILL
                r += 1

                for ci, h in enumerate(headers, 1):
                    cell = ws.cell(row=r, column=ci, value=h)
                    cell.font = HEADER_FONT
                    cell.fill = HEADER_FILL
                    cell.alignment = Alignment(horizontal="center")
                    cell.border = THIN_BORDER
                r += 1

                for row_data in rows:
                    for ci, val in enumerate(row_data, 1):
                        cell = ws.cell(row=r, column=ci, value=val)
                        cell.border = THIN_BORDER
                        cell.alignment = Alignment(wrap_text=True, vertical="top")
                    r += 1

                r += 1

                if col_widths:
                    for ci, w in enumerate(col_widths, 1):
                        ws.column_dimensions[get_column_letter(ci)].width = w

            write_section(
                "基本情報",
                ["項目", "内容"],
                [
                    ["タイトル", task.title],
                    ["ステータス", f"{STATUS_ICONS_DICT.get(task.status, '')} {task.status}"],
                    ["優先度", f"{PRIORITY_ICONS_DICT.get(task.priority, '')} {task.priority}"],
                    ["タイプ", task.task_type],
                    ["担当者", task.owner],
                    ["依頼者", task.requester],
                    ["開始日", task.start_date],
                    ["期限", task.due_date],
                    ["作成日", task.created_at],
                    ["更新日", task.updated_at],
                ],
                col_widths=[18, 50],
            )

            if task.core_objective:
                write_section(
                    "コア目標",
                    ["内容"],
                    [[task.core_objective]],
                    col_widths=[68],
                )

            if task.description:
                write_section(
                    "説明",
                    ["内容"],
                    [[task.description]],
                    col_widths=[68],
                )

            if task.wbs_items:
                wbs_rows = []
                total_est = 0
                total_act = 0
                for item in task.wbs_items:
                    wbs_rows.append([
                        item.phase,
                        item.task_name,
                        item.estimated_hours or 0,
                        item.actual_hours or 0,
                        item.progress,
                    ])
                    total_est += item.estimated_hours or 0
                    total_act += item.actual_hours or 0
                wbs_rows.append(["合計", "", total_est, total_act, ""])
                write_section(
                    "WBS 工数管理",
                    ["フェーズ", "タスク", "見積(h)", "実績(h)", "進捗"],
                    wbs_rows,
                    col_widths=[18, 28, 10, 10, 14],
                )

            risk_rows = []
            for r_item in task.risks:
                desc = r_item.description if isinstance(r_item, RiskItem) else str(r_item)
                mit = r_item.mitigation if isinstance(r_item, RiskItem) and r_item.mitigation else ""
                risk_rows.append([desc, mit])
            if task.external_dependencies:
                for dep in task.external_dependencies:
                    risk_rows.append([f"依存: {dep}", ""])
            if task.buffer_hours:
                risk_rows.append([f"バッファ時間: {task.buffer_hours}h", ""])

            if risk_rows:
                write_section(
                    "リスク・依存",
                    ["内容", "対応策"],
                    risk_rows,
                    col_widths=[34, 34],
                )

            log_rows = []
            for log in task.daily_logs:
                if isinstance(log, DailyLogItem):
                    log_rows.append([f"📅 {log.date}", ""])
                    for item in log.items:
                        log_rows.append(["", f"- {item}"])
                    if log.notes:
                        log_rows.append(["", f"  ノート: {log.notes}"])
                    log_rows.append(["", ""])
            if log_rows:
                write_section(
                    "Daily Log",
                    ["日付", "内容"],
                    log_rows,
                    col_widths=[18, 50],
                )

            hist_rows = []
            for h in task.change_history:
                item = h if isinstance(h, dict) else h.__dict__
                hist_rows.append([
                    item.get("date", ""),
                    item.get("change", ""),
                    item.get("changed_by", ""),
                ])
            if hist_rows:
                write_section(
                    "変更履歴",
                    ["日付", "変更内容", "変更者"],
                    hist_rows,
                    col_widths=[14, 40, 14],
                )

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        body = output.read()
        filename = f"task_board_{datetime.now().strftime('%Y%m%d')}.xlsx"
        self.send_response(200)
        self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def _send_json(self, data: dict, status: int = 200):
        body = json.dumps(data, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def _send_html(self, html: str, status: int = 200):
        body = html.encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def _read_body(self) -> dict:
        length = int(self.headers.get("Content-Length", 0))
        if length == 0:
            return {}
        raw = self.rfile.read(length)
        return json.loads(raw.decode("utf-8"))

    def _parse_path(self) -> tuple:
        parsed = re.match(r"^/api/tasks(?:/([a-zA-Z0-9_-]+))?/?$", self.path)
        if parsed:
            return ("api", parsed.group(1))
        if self.path == "/api/export/excel":
            return ("export_excel", None)
        if self.path == "/" or self.path == "":
            return ("index", None)
        return ("unknown", None)

    def do_GET(self):
        kind, task_id = self._parse_path()

        if kind == "index":
            index_path = os.path.join(TEMPLATE_DIR, "index.html")
            if os.path.exists(index_path):
                with open(index_path, "r", encoding="utf-8") as f:
                    self._send_html(f.read())
            else:
                self._send_json({"error": "template not found"}, 500)
            return

        if kind == "api":
            if task_id:
                task = get_task(task_id)
                if not task:
                    self._send_json({"error": "not found"}, 404)
                    return
                self._send_json(_task_to_dict(task))
            else:
                tasks = load_tasks()
                self._send_json([_task_to_dict(t) for t in tasks])
            return

        if kind == "export_excel":
            self._send_excel()
            return

        self._send_json({"error": "not found"}, 404)

    def do_POST(self):
        kind, task_id = self._parse_path()

        if kind == "api" and task_id is None:
            data = self._read_body()
            task = _dict_to_task(data)
            created = create_task(task)
            self._send_json(_task_to_dict(created), 201)
            return

        self._send_json({"error": "not found"}, 404)

    def do_PUT(self):
        kind, task_id = self._parse_path()

        if kind == "api" and task_id:
            task = get_task(task_id)
            if not task:
                self._send_json({"error": "not found"}, 404)
                return
            data = self._read_body()
            data["id"] = task_id
            updated = _dict_to_task(data)
            result = update_task(updated)
            if not result:
                self._send_json({"error": "update failed"}, 500)
                return
            self._send_json(_task_to_dict(result))
            return

        self._send_json({"error": "not found"}, 404)

    def do_DELETE(self):
        kind, task_id = self._parse_path()

        if kind == "api" and task_id:
            ok = delete_task(task_id)
            if not ok:
                self._send_json({"error": "not found"}, 404)
                return
            self._send_json({"ok": True})
            return

        self._send_json({"error": "not found"}, 404)

    def log_message(self, format, *args):
        print(f"[{datetime.now().strftime('%H:%M:%S')}] {args[0]} {args[1]} {args[2]}")


def run():
    server = HTTPServer((HOST, PORT), TaskHandler)
    print(f"Server running at http://{HOST}:{PORT}")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\nShutting down...")
        server.server_close()


if __name__ == "__main__":
    run()
