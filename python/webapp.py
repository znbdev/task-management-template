import io
from datetime import datetime

from flask import Flask, jsonify, request, render_template, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from models import Task, WBSItem, DailyLogItem, RiskItem, ChangeHistoryItem
from models import STATUSES, PRIORITIES, TASK_TYPES, PHASES
from storage import load_tasks, save_tasks, create_task, update_task, delete_task, get_task

app = Flask(__name__)


def _task_to_dict(t: Task) -> dict:
    d = t.to_dict()
    d["total_estimated"] = t.total_estimated
    d["total_actual"] = t.total_actual
    d["status_icon"] = t.status_icon
    d["priority_icon"] = t.priority_icon
    return d


def _dict_to_task(data: dict) -> Task:
    def _parse_list(key, cls):
        items = data.get(key, [])
        return [cls(**item) if isinstance(item, dict) else item for item in items]

    t = Task(
        id=data.get("id", ""),
        title=data.get("title", ""),
        description=data.get("description", ""),
        status=data.get("status", STATUSES[0]),
        priority=data.get("priority", PRIORITIES[2]),
        task_type=data.get("task_type", TASK_TYPES[0]),
        owner=data.get("owner", ""),
        requester=data.get("requester", ""),
        start_date=data.get("start_date", ""),
        due_date=data.get("due_date", ""),
        core_objective=data.get("core_objective", ""),
        deliverables=data.get("deliverables", []),
        success_criteria=data.get("success_criteria", []),
        wbs_items=_parse_list("wbs_items", WBSItem),
        risks=[RiskItem(**r) if isinstance(r, dict) else r for r in data.get("risks", [])],
        external_dependencies=data.get("external_dependencies", []),
        buffer_hours=float(data.get("buffer_hours", 0)),
        daily_logs=_parse_list("daily_logs", DailyLogItem),
        resources=data.get("resources", []),
        change_history=[ChangeHistoryItem(**h) if isinstance(h, dict) else h for h in data.get("change_history", [])],
        tags=data.get("tags", []),
        created_at=data.get("created_at", ""),
        updated_at=data.get("updated_at", ""),
    )
    return t


@app.route("/")
def index():
    return render_template("index.html",
                           statuses=STATUSES,
                           status_icons=["🟩", "🟨", "🟦", "⬜", "❌"],
                           priorities=PRIORITIES,
                           task_types=TASK_TYPES,
                           phases=PHASES)


@app.route("/api/tasks", methods=["GET"])
def api_list_tasks():
    tasks = load_tasks()
    return jsonify([_task_to_dict(t) for t in tasks])


@app.route("/api/tasks", methods=["POST"])
def api_create_task():
    data = request.get_json(force=True)
    task = _dict_to_task(data)
    created = create_task(task)
    return jsonify(_task_to_dict(created)), 201


@app.route("/api/tasks/<task_id>", methods=["GET"])
def api_get_task(task_id):
    task = get_task(task_id)
    if not task:
        return jsonify({"error": "not found"}), 404
    return jsonify(_task_to_dict(task))


@app.route("/api/tasks/<task_id>", methods=["PUT"])
def api_update_task(task_id):
    task = get_task(task_id)
    if not task:
        return jsonify({"error": "not found"}), 404
    data = request.get_json(force=True)
    data["id"] = task_id
    updated = _dict_to_task(data)
    result = update_task(updated)
    if not result:
        return jsonify({"error": "update failed"}), 500
    return jsonify(_task_to_dict(result))


@app.route("/api/tasks/<task_id>", methods=["DELETE"])
def api_delete_task(task_id):
    ok = delete_task(task_id)
    if not ok:
        return jsonify({"error": "not found"}), 404
    return jsonify({"ok": True})


@app.route("/api/export/excel")
def api_export_excel():
    tasks = load_tasks()
    wb = Workbook()
    wb.remove(wb.active)

    HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
    HEADER_FILL = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    SECTION_FONT = Font(bold=True, size=11, color="2C3E50")
    SECTION_FILL = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")
    THIN_BORDER = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    STATUS_ICONS_DICT = dict(zip(STATUSES, ["🟩", "🟨", "🟦", "⬜", "❌"]))
    PRIORITY_ICONS_DICT = dict(zip(PRIORITIES, ["🔴", "🟠", "🟡", "🔵"]))

    for task in tasks:
        sheet_name = task.title[:31] if task.title else "無題"
        ws = wb.create_sheet(title=sheet_name)

        r = 1

        def write_section(title, headers, rows, col_widths=None):
            nonlocal r
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
            cell = ws.cell(row=r, column=1, value=title)
            cell.font = SECTION_FONT
            cell.fill = SECTION_FILL
            r += 1

            for ci, h in enumerate(headers, 1):
                cell = ws.cell(row=r, column=ci, value=h)
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = Alignment(horizontal="center")
                cell.border = THIN_BORDER
            r += 1

            for row_data in rows:
                for ci, val in enumerate(row_data, 1):
                    cell = ws.cell(row=r, column=ci, value=val)
                    cell.border = THIN_BORDER
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                r += 1

            r += 1

            if col_widths:
                for ci, w in enumerate(col_widths, 1):
                    ws.column_dimensions[get_column_letter(ci)].width = w

        # Basic Info
        write_section(
            "📄 基本情報",
            ["項目", "内容"],
            [
                ["タイトル", task.title],
                ["ステータス", f"{STATUS_ICONS_DICT.get(task.status, '')} {task.status}"],
                ["優先度", f"{PRIORITY_ICONS_DICT.get(task.priority, '')} {task.priority}"],
                ["タイプ", task.task_type],
                ["担当者", task.owner],
                ["依頼者", task.requester],
                ["開始日", task.start_date],
                ["期限", task.due_date],
                ["作成日", task.created_at],
                ["更新日", task.updated_at],
            ],
            col_widths=[18, 50],
        )

        # Objective
        if task.core_objective:
            write_section(
                "🎯 コア目標",
                ["内容"],
                [[task.core_objective]],
                col_widths=[68],
            )

        # Description
        if task.description:
            write_section(
                "📝 説明",
                ["内容"],
                [[task.description]],
                col_widths=[68],
            )

        # WBS
        if task.wbs_items:
            wbs_rows = []
            total_est = 0
            total_act = 0
            for item in task.wbs_items:
                wbs_rows.append([
                    item.phase,
                    item.task_name,
                    item.estimated_hours or 0,
                    item.actual_hours or 0,
                    item.progress,
                ])
                total_est += item.estimated_hours or 0
                total_act += item.actual_hours or 0
            wbs_rows.append(["📊 合計", "", total_est, total_act, ""])
            write_section(
                "⏳ WBS 工数管理",
                ["フェーズ", "タスク", "見積(h)", "実績(h)", "進捗"],
                wbs_rows,
                col_widths=[18, 28, 10, 10, 14],
            )

        # Risks & Dependencies
        risk_rows = []
        for r_item in task.risks:
            desc = r_item.description if isinstance(r_item, RiskItem) else str(r_item)
            mit = r_item.mitigation if isinstance(r_item, RiskItem) and r_item.mitigation else ""
            risk_rows.append([desc, mit])
        if task.external_dependencies:
            for dep in task.external_dependencies:
                risk_rows.append([f"依存: {dep}", ""])
        if task.buffer_hours:
            risk_rows.append([f"バッファ時間: {task.buffer_hours}h", ""])

        if risk_rows:
            write_section(
                "⚠️ リスク・依存",
                ["内容", "対応策"],
                risk_rows,
                col_widths=[34, 34],
            )

        # Daily Log
        log_rows = []
        for log in task.daily_logs:
            if isinstance(log, DailyLogItem):
                log_rows.append([f"📅 {log.date}", ""])
                for item in log.items:
                    log_rows.append(["", f"- {item}"])
                if log.notes:
                    log_rows.append(["", f"  ノート: {log.notes}"])
                log_rows.append(["", ""])
        if log_rows:
            write_section(
                "📝 Daily Log",
                ["日付", "内容"],
                log_rows,
                col_widths=[18, 50],
            )

        # Change History
        hist_rows = []
        for h in task.change_history:
            item = h if isinstance(h, dict) else h.__dict__
            hist_rows.append([
                item.get("date", ""),
                item.get("change", ""),
                item.get("changed_by", ""),
            ])
        if hist_rows:
            write_section(
                "🔄 変更履歴",
                ["日付", "変更内容", "変更者"],
                hist_rows,
                col_widths=[14, 40, 14],
            )

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"task_board_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(output, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
